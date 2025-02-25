import streamlit as st
import pandas as pd
import openpyxl
import io
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from huggingface_hub import InferenceClient
from transformers import AutoTokenizer, AutoModelForCausalLM, pipeline

# Set up Hugging Face API Key
HF_API_KEY = "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 "
client = InferenceClient(provider="together", api_key=HF_API_KEY)

# Load Local Model
MODEL_NAME = "mistralai/Mistral-7B-Instruct-v0.3"
tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
model = AutoModelForCausalLM.from_pretrained(MODEL_NAME)
pipe = pipeline("text-generation", model=model)

# Initialize session state for uploaded files
if "uploaded_files" not in st.session_state:
    st.session_state.uploaded_files = {}

# Streamlit UI Setup
st.set_page_config(page_title="AI-Powered Timetable", layout="wide")
st.markdown("<h1 style='text-align: center; color: #4CAF50;'>📅 AI-Powered Timetable</h1>", unsafe_allow_html=True)

# File Upload Section
st.sidebar.markdown("## 📂 Upload Your Timetable Files")
uploaded_master = st.sidebar.file_uploader("Upload Master Timetable", type=["xlsx"])
uploaded_lab = st.sidebar.file_uploader("Upload Lab Timetable", type=["xlsx"])
uploaded_classroom = st.sidebar.file_uploader("Upload Classroom Timetable", type=["xlsx"])
uploaded_individual = st.sidebar.file_uploader("Upload Individual Timetable", type=["xlsx"])

# Save uploaded files locally
if uploaded_master:
    st.session_state.uploaded_files["Master Timetable"] = uploaded_master
if uploaded_lab:
    st.session_state.uploaded_files["Lab Timetable"] = uploaded_lab
if uploaded_classroom:
    st.session_state.uploaded_files["Classroom Timetable"] = uploaded_classroom
if uploaded_individual:
    st.session_state.uploaded_files["Individual Timetable"] = uploaded_individual

# Define paths for uploaded files
TIMETABLE_FILES = {name: file for name, file in st.session_state.uploaded_files.items()}

# Load Timetable Data
def load_timetable(sheet_name):
    if sheet_name not in TIMETABLE_FILES:
        return None
    file = TIMETABLE_FILES[sheet_name]
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    return [row for row in sheet.iter_rows(values_only=True)]

# Convert Timetable to AI-Readable Format
def format_timetable(sheet_name):
    data = load_timetable(sheet_name)
    if data is None:
        return "No data available."
    formatted_data = "\n".join([", ".join(map(str, row)) for row in data[1:]])
    return f"Timetable Data:\n{formatted_data}"

# Ask Mistral AI a question using Hugging Face Inference API
def ask_mistral_api(query):
    messages = [{"role": "user", "content": query}]
    completion = client.chat.completions.create(
        model=MODEL_NAME,
        messages=messages,
        max_tokens=500,
    )
    return completion.choices[0].message["content"]

# Ask Mistral AI a question using local model
def ask_mistral_local(query):
    inputs = tokenizer(query, return_tensors="pt")
    outputs = model.generate(**inputs, max_new_tokens=200)
    response = tokenizer.decode(outputs[0], skip_special_tokens=True)
    return response

# Auto-Schedule Missing Slots
def auto_schedule(sheet_name):
    if sheet_name not in TIMETABLE_FILES:
        return "No timetable uploaded."

    file = TIMETABLE_FILES[sheet_name]
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    # Find empty slots
    empty_slots = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if None in row or "" in row:
            empty_slots.append(row_idx)

    # Ask AI for suggestions
    for row_idx in empty_slots:
        query = f"Suggest a subject and faculty for the empty slot in row {row_idx}."
        suggestion = ask_mistral_local(query)  # Use local model

        try:
            subject, faculty = suggestion.split(", Faculty: ")
            subject = subject.replace("Subject: ", "").strip()
            faculty = faculty.strip()
            sheet.cell(row=row_idx, column=4, value=subject)
            sheet.cell(row=row_idx, column=5, value=faculty)
        except:
            continue  

    wb.save(file.name)
    return f"Auto-scheduling completed for {len(empty_slots)} slots."

# Timetable Selection
timetable_type = st.sidebar.selectbox("📌 Select Timetable:", list(TIMETABLE_FILES.keys()) if TIMETABLE_FILES else ["No files uploaded"])

# Load & Display Timetable
if st.sidebar.button("📂 Load Timetable"):
    if timetable_type not in TIMETABLE_FILES:
        st.error("No file uploaded for this timetable!")
    else:
        file = TIMETABLE_FILES[timetable_type]
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        st.write(f"### {timetable_type} 📋")
        st.dataframe(df, use_container_width=True)

        # Download as Excel
        df.to_excel("Updated_Timetable.xlsx", index=False)
        st.download_button("📥 Download Updated Excel", data=open("Updated_Timetable.xlsx", "rb"), file_name="Updated_Timetable.xlsx")

        # Generate PDF Function
        def generate_pdf(dataframe):
            pdf_buffer = io.BytesIO()
            c = canvas.Canvas(pdf_buffer, pagesize=letter)
            c.drawString(100, 750, f"{timetable_type} Timetable")
            y = 720
            for index, row in dataframe.iterrows():
                c.drawString(100, y, str(row.values))
                y -= 20
                if y < 50:
                    c.showPage()
                    y = 750
            c.save()
            pdf_buffer.seek(0)
            return pdf_buffer

        # Button to Generate PDF
        pdf_data = generate_pdf(df)
        st.download_button("📄 Download as PDF", data=pdf_data, file_name=f"{timetable_type}_Timetable.pdf", mime="application/pdf")

# AI Query Section
st.markdown("## 🤖 Ask Mistral AI About Your Timetable")
user_query = st.text_input("Type your question here (e.g., 'Who is free at 10 AM on Monday?')")

if st.button("Ask AI via API"):
    if timetable_type not in TIMETABLE_FILES:
        st.error("No file uploaded for AI to analyze!")
    else:
        ai_response = ask_mistral_api(user_query)
        st.write("🧠 **Mistral AI Suggests:**", ai_response)

if st.button("Ask AI via Local Model"):
    ai_response = ask_mistral_local(user_query)
    st.write("🧠 **Mistral AI Suggests:**", ai_response)

# Auto-Schedule Button
if st.sidebar.button("⚡ Auto-Schedule with AI"):
    if timetable_type not in TIMETABLE_FILES:
        st.error("No file uploaded for auto-scheduling!")
    else:
        result = auto_schedule(timetable_type)
        st.success(result)

# Footer
st.markdown("<hr style='border:1px solid #4CAF50;'>", unsafe_allow_html=True)
st.markdown("<p style='text-align: center;'>Developed with 🚀 by ChatGPT</p>", unsafe_allow_html=True)
